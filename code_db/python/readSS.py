import openpyxl
import groq_first_pass
from openpyxl.utils import column_index_from_string
import time
finalJson_ = dict()

def read_excel_file(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook( file_path, read_only=True )
    # Get the specified sheet in the workbook
    localD = dict()

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        time.sleep( 1 )
        print('Iterating over sheet->', file_path, sheet_name)
        
        num_rows_to_consider_ , frame_ = 4, ''
        try:
            for rowidx, row in enumerate( sheet.iter_rows(values_only=True) ):

                if rowidx > num_rows_to_consider_: break
                for cell in row:
                    frame_ += str(cell) + '\t'

                frame_ += '\n'

            print('Sending to LLM for summary->', frame_)

            summary_ = groq_first_pass.returnLLMResponse( frame_ )
            ## append file name, sheet name
            localD[ sheet_name ] = file_path.split('/')[-1] + ' ' + sheet_name + ' ' + summary_
        except:
            print( 'EXCPN-> '+file_path + ' ' + sheet_name + ' ' + traceback.format_exc() )

    finalJson_[ file_path.split('/')[-1] ] = localD


if __name__ == "__main__":
    # Example usage:
    import os, json
    file_ll_ = os.listdir( './DATA' )

    for file_path in file_ll_:
        print('ITERATING ->', file_path)
        read_excel_file( './DATA/' + file_path )

    with open( 'SpreadSheetSummary.json', 'a' ) as fp:
        json.dump( finalJson_, fp )
