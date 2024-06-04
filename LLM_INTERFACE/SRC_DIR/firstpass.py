from openpyxl import load_workbook

def find_table_bounds(sheet):
    max_row = sheet.max_row
    max_col = sheet.max_column

    # Initialize variables to track the bounds
    start_row ,end_row ,start_col ,end_col = None, None, None, None

    # Iterate over rows to find the start and end rows
    for row in range(1, max_row + 1):
        if all(sheet.cell(row=row, column=col).value is None for col in range(1, max_col + 1)):
            if start_row is None:
                continue  # Skip empty rows before the table
            else:
                end_row = row - 1
                break
        elif start_row is None:
            start_row = row
    
    # Iterate over columns to find the start and end columns
    for col in range(1, max_col + 1):
        #for row in range(start_row, end_row):
        #    print('ROW NUM->', col, ' VALUE: ', sheet.cell(row=row, column=col).value )

        if all(sheet.cell(row=row, column=col).value is None for row in range(start_row, end_row)):
            if start_col is None:
                continue  # Skip empty columns before the table
            else:
                end_col = col - 1
                break
        elif start_col is None:
            start_col = col

    return start_row, end_row, start_col, end_col

# Example usage
workbook = load_workbook( './DATA/Mark Training samples % .xlsx' )
sheet = workbook.active

start_row, end_row, start_col, end_col = find_table_bounds(sheet)
print("Table bounds: Start Row =", start_row, ", End Row =", end_row, ", Start Column =", start_col, ", End Column =", end_col, sheet)

