import json, math, sys, traceback, copy, multiprocessing, os
from dateutil import parser
import numpy as np
import openpyxl
from openpyxl.utils import column_index_from_string
import time, random, datetime
import pandas as pd
from sklearn.decomposition import PCA

import query_llama3_via_groq as llama3
import query_gpt_via_groq as openai
import createJsonFeats
import db_utils

def is_date( input_str):
        ## first check for INT and FLOAT since parser foolishly accepts ints
        try:
            _ = int( input_str )
            return None
        except:
            pass

        try:
            _ = float( input_str )
            return None
        except:
            pass

        try:
            return parser.parse(input_str)
        except ValueError:
            return None

def process( colNum, sheet, tbl ):
        dt_counts_ = []

        for rw in range( tbl['START_ROW'], tbl['END_ROW'] ):
                dtVal_ = is_date( str( sheet.cell(row=rw, column=colNum).value ) )

                if dtVal_ is not None : 
                    dt_counts_.append( dtVal_ ) 

        if len( dt_counts_ ) >= ( tbl['END_ROW'] - tbl['START_ROW'] )/2: 
                ## defensive chk to ensure dt counts are high
                print('Dt Col found !', colNum)
                ## sort the values to get range
                sorted_dates_ = sorted( dt_counts_ )
                print('Dt range->', sorted_dates_[0], sorted_dates_[-1] )

                return ( True, sorted_dates_[0].strftime('%B %d, %Y'), sorted_dates_[-1].strftime('%B %d, %Y') )

        return ( False, None, None )

class GenerateXLMetaInfo:
    def __init__(self, file_path, llm='LLAMA'):
        """
        Initialize the GenerateXmlMetaInfo class with the XML file.

        Parameters:
        - xml_file (str): The path to the XML file.
        """
        self.file_path = file_path
        self.masterInfo_ = dict()
        self.llm_framework_ = llm
        self.sheet = None
        self.sklearn_pca_object_ = PCA()
        self.add_ai_summary_to_embedding_ = True
        self.chunk_size_ = 500 ## approx 1024 tokens

        self.sz_of_phrase_, self.unique_str_thresh_, self.number_str_thresh_ = 5, 0.5, 0.4
        self.pca_var_min_contrib, self.feature_contribution_per_thresh_ = 0.5, 0.3
        self.max_tables_per_sheet_ = 10 ## at times a single sheet can have multiple tables
        self.num_rows_to_consider_, self.col_thresh_, self.minElemsInTable, self.max_rows_variance = 4, 0.8, 6, 100
        self.default_max_col_ , self.default_max_row_, self.max_cols_for_review_, \
                self.min_num_distinct_values_, self.max_elements_for_summary_ = 50, 50, 10, 3, 15

        if llm == 'OPENAI':
            self.query_fn_ = openai
        else:
            ## default , llama3 inferred via groq
            self.query_fn_ = llama3
   
    def find_bounds( self, sheet, max_row, max_col, start_row ,end_row ,start_col ,end_col ):

        # Iterate over rows to find the start and end rows
        start_row_idx_ = 1 if start_row is None else start_row
        start_col_idx_ = 1 if start_col is None else start_col
        
        ## need to add 2 to max rw and col since max_row of sheet returns the last row with data
        ## and range will stop at max_row - 1 
        for row in range( start_row_idx_ , max_row + 2):
            if all( self.sheet.cell(row=row, column=col).value is None for col in range(1, max_col + 2)):
                if start_row is None:
                    continue  # Skip empty rows before the table
                else:
                    end_row = row - 1
                    break
            elif start_row is None:
                start_row = row
           
        # Iterate over columns to find the start and end columns
        for col in range( start_col_idx_, max_col + 2):
            #for row in range(start_row, end_row):
            #    print('ROW NUM->', col, ' VALUE: ', sheet.cell(row=row, column=col).value )
            if end_row is None: continue

            if all( self.sheet.cell(row=row, column=col).value is None for row in range(start_row, end_row)):
                if start_col is None:
                    continue  # Skip empty columns before the table
                else:
                    end_col = col - 1
                    break
            elif start_col is None:
                start_col = col
            

        #print('Found tables between-> start_row, end_row, start_col, end_col = ',\
        #        start_row, end_row, start_col, end_col )

        return start_row, end_row, start_col, end_col

    def is_hdr_row_format( self, tbl_bound, sheet ):
        
        num_str_cols_ = 0
        for col_ctr in range( tbl_bound['START_COL'], tbl_bound['END_COL'] ):
            if type( self.sheet.cell(row=tbl_bound['START_ROW'], column=col_ctr).value ) == str:
                num_str_cols_ += 1

        if num_str_cols_ < ( tbl_bound['END_COL'] - tbl_bound['START_COL'] ): return False

        return True

    def find_tables(self, sheet):
        ## NOTE -> sheet.max_row and sheet.max_column is NOT WORKING !! NEED TO FIX
        ## default is stop gap
        max_row = sheet.max_row if sheet.max_row is not None else self.default_max_row_
        max_col = sheet.max_column if sheet.max_column is not None else self.default_max_col_
        table_bounds_ = []

        print('KKR->', max_row, max_col)
        timer_ = time.time()
        # Initialize variables to track the bounds
        start_row ,end_row ,start_col ,end_col = None, None, 1, sheet.max_column

        ## do a first pass to find the first table
        start_row, end_row, start_col, end_col = self.find_bounds( sheet, max_row, max_col, start_row ,\
                                                                         end_row ,start_col ,end_col )

        print('Time taken->', time.time() - timer_, start_row, end_row, start_col, end_col)    
        init_end_col = copy.copy( end_col )

        table_bounds_.append( { 'START_ROW': start_row, 'END_ROW': end_row,\
                                    'START_COL': start_col, 'END_COL': end_col } ) 

        ## now iterate from end_row to max_row to find all tables row wise
        while end_row is not None:
            end_row += 2 ## increment by 2 since we need to look ahead and see if any more tables exist !
            ##              if u increment by 1 then u will end up on the same blank line that stopped the prev tbl  
            ## start_row is assigned the value of end_row from above and end_row is made None
            if end_row >= max_row: break

            #print('DUM ROW->', end_row)
            start_row, end_row, start_col, end_col = self.find_bounds( sheet, max_row, max_col, end_row ,\
                                                                             None , None , None )

            if ( start_col is None or end_col is None ) or \
                    ( abs( start_row - end_row )*abs( start_col - end_col ) ) <= self.minElemsInTable: continue    

            table_bounds_.append( { 'START_ROW': start_row, 'END_ROW': end_row,\
                                    'START_COL': start_col, 'END_COL': end_col } ) 
        
        ## now iterate from end_col to max_col to find all tables cols wise
        while init_end_col is not None:
            init_end_col += 2 ## increment by 1 since we need to look ahead and see if any more tables exist !
            ## start_row is assigned the value of end_row from above and end_row is made None
            if init_end_col >= max_col: break

            #print('DUM COL->', init_end_col)
            start_row, end_row, start_col, end_col = self.find_bounds( sheet, max_row, max_col, None ,\
                                                                             None , init_end_col , None )

            if ( start_col >= end_col ): continue    

            table_bounds_.append( { 'START_ROW': start_row, 'END_ROW': end_row,\
                                    'START_COL': start_col, 'END_COL': end_col } ) 

        ## init star and end col to min and max
        for tab in table_bounds_: tab['START_COL'] = 1; tab['END_COL'] = max_col;
        ## remove dupes
        tmp_, dupe = [], set()

        for idx1, tab1 in enumerate( table_bounds_ ):
            for idx2, tab2 in enumerate( table_bounds_ ):
                if idx1 <= idx2: continue
                if tab2['START_ROW'] >= tab1['START_ROW'] and tab2['END_ROW'] <= tab1['END_ROW']:
                    dupe.add( idx2 )

        for idx, tab in enumerate( table_bounds_ ):
            if idx not in list( dupe ):
                tmp_.append( tab )

        ## blend tables - in case the rows are FPs
        final_resp_ = []
        if len( tmp_ ) > 1:
            last_tbl_ = tmp_[0]
            final_resp_.append( last_tbl_ )
            ## check if the first row is not all STR
            for ctr in range( 1, len( tmp_ ) ):
                if self.is_hdr_row_format( tmp_[ctr], sheet ) == False:
                    ## blend with the last table
                    final_resp_[-1]['END_ROW'] = tmp_[ctr]['END_ROW']
                else:
                    final_resp_.append( table_bounds_ )
        else:
            final_resp_ = tmp_

        return final_resp_[ : min( self.max_tables_per_sheet_, len( final_resp_ ) ) ]

    def findDateRange( self, tbl ):

        colRange_ = list( range( tbl['START_COL'], tbl['END_COL'] ) )

        for col in colRange_:
            ## process was taken out of the class only because multi processing map refused to pickle
            ## a method that was part of the class ..and it proved way slower ..so parallels been removed..lol
            results = process(col, self.sheet, tbl)
            if results[0] is True:
                    return str( results[1] ) +' To '+ str( results[2] )

        return (None, None)

    def findHeaderInfo(self, tbl):
        """
        Find header information from the XL file.
        take the first 2 rows and then to be on the safe side also take the 
        first 2 columns ( in case the col headers are just numbers / % etc and the row contain item name in the first col )
        send it to the LLM for a summary
        ALSO there's no need to take all ROWS and COLS .. some 10-15 elements are more than enough but can be adjustedfor domains that need more
        """

        hdr_row_start_ = self.findHdrRow( tbl )
        row_starter_ = tbl['START_ROW'] if hdr_row_start_ is None else hdr_row_start_

        col_frame_ = ''

        for rw in range( row_starter_ , min( row_starter_ + self.num_rows_to_consider_, tbl['END_ROW'] ) ):
            for col in range( tbl['START_COL'], min( self.max_elements_for_summary_, tbl['END_COL'] + 1 ) ):
                col_frame_ += '\t' + str( self.sheet.cell(row=rw, column=col).value )

            col_frame_ += '\n'

        return col_frame_

    def findHighVarianceColumns(self, start_hdr_row_, sheet, tbl ):
        '''
        iterate through columns that have numeric values and figure out the more important columns
        num of rows - we can restrict it to lets say 1k rows ..should suffice 
        '''
        numeric_frame_, high_var_indices_, hdr_col_names_ = dict(), set(), []
        end_row_ =  min( tbl['START_ROW'] + self.max_rows_variance , tbl['END_ROW'] + 1 )
        ## add 1 to the start row since we dont want to include the header value
        start_row_ = ( tbl['START_ROW'] if start_hdr_row_ is None else start_hdr_row_ ) + 1 
        
        print('BIGGIE-> start_row_, end_row_ = ', start_row_, end_row_)

        for col_ctr in range( tbl['START_COL'], tbl['END_COL']+1 ):
                hdr_col_names_.append( str( self.sheet.cell(row=start_row_-1, column=col_ctr).value ) )

        try:
            for col_ctr in range( tbl['START_COL'], tbl['END_COL']+1 ):
                col_arr_ = [ 'NA' for x in range( ( end_row_ - start_row_ ) + 1 ) ]

                for idx, row_ctr in enumerate( range( start_row_, end_row_ ) ):
                    col_arr_[ idx ] = str( self.sheet.cell(row=row_ctr, column=col_ctr).value )

                ## standardize the column since PCA better be done on std values
                col_set_ = set( col_arr_ )
                ## convert the variables into unique IDs
                uid = [ list( col_set_ ).index( x ) for x in col_arr_ ]
                max_uid_ = np.max( uid )
                ## normalize the int values
                numeric_frame_[ col_ctr ] = [ x/max_uid_ for x in uid ]

            if len( numeric_frame_.keys() ) > 0:
                ## now transpose the contents of the frame since we want it to retain the shape of a column
                transposed_ = np.transpose( np.asarray( list( numeric_frame_.values() ) ) )
                #print('The val of transposed_->', transposed_)
                ## perform PCA and pick the most high variance columns
                ## the number of components to be picked will be decided by the thresh self.percent_pca_var_
                self.sklearn_pca_object_.fit( transposed_ )
                ## components_loading_ will give you principal component wise contribution of the features
                components_loading_ = self.sklearn_pca_object_.components_
                ## only consider those components that contribute to 90% or whatever threshold level of variance
                relevant_loading_ = components_loading_[0] \
                                    if self.sklearn_pca_object_.explained_variance_ratio_[0] > self.pca_var_min_contrib \
                                    else []

                #print('LOADING AND REL_LOADING->', components_loading_, relevant_loading_)
                key_list_ = list( numeric_frame_.keys() )

                for feat_idx, feat_contribution in enumerate( relevant_loading_ ):
                        if feat_contribution >= self.feature_contribution_per_thresh_: 

                            high_var_indices_.add( hdr_col_names_[ key_list_[ feat_idx ] ] )

                            #print('Adding ', hdr_col_names_[ key_list_[ feat_idx ] ],' As a high variance col')
        except:
            pass

        return list( high_var_indices_ ), hdr_col_names_


    def returnSummary(self, tbl ):
        '''
        take the first few rows to try and generate a coherent summary for the type of the data present
        i am also considering transposing the first few rows to see how different the summary looks
        ALSO maybe limiting the number of columns makes sense
        '''
        frame_, transposed_frame_, start_hdr_row_ = '', '', self.findHdrRow( tbl )
        
        time_ = time.time()
        high_variance_cols_, hdr_col_names_ = self.findHighVarianceColumns( start_hdr_row_, self.sheet, tbl )
        print('Time taken to find high var cols ->', time.time() - time_)
        print('AND THEY ARE->', high_variance_cols_)

        frame_num_contours_, transposed_frame_contours_ = 0, 0
        ## NATURAL order -> left to right, top to bottom
        for row_ctr in range( tbl['START_ROW'] if start_hdr_row_ is None else start_hdr_row_\
                              , min( tbl['START_ROW']+self.num_rows_to_consider_ , tbl['END_ROW']+1 ) ):

            for col_ctr in range( tbl['START_COL'], min( self.max_elements_for_summary_, tbl['END_COL']+1 ) ):

                frame_ += '\t' + str( self.sheet.cell(row=row_ctr, column=col_ctr).value )
                frame_num_contours_ += 1

            frame_ += '\n'

        return frame_, high_variance_cols_, list( set(hdr_col_names_) )

    def findHdrRow( self, tbl ):

        total_cols_ = tbl['END_COL'] - tbl['START_COL']

        for row_ctr in range( tbl['START_ROW'], \
                              min( tbl['START_ROW']+self.num_rows_to_consider_ , tbl['END_ROW']+1 ) ):
            num_non_blank_ = 0
            
            for col_ctr in range( tbl['START_COL'], tbl['END_COL'] ):
                if self.sheet.cell(row=row_ctr, column=col_ctr).value is not None and \
                        len( str( self.sheet.cell(row=row_ctr, column=col_ctr).value ) ) > 0: 
                    num_non_blank_ += 1

            ## only if the number of hdr columns is in the ballpark w.r.t. total number of columns
            ## should we start the table ..at times the header table is split across more than 1 row
            if total_cols_ > 1 and (num_non_blank_/total_cols_) > self.col_thresh_:
                return row_ctr

        return None # so default value of row #1 applies to table start
    
    def createDBRec( self, summary_D, mode='NORM' ):

        insertRec = dict()
        insertRec['docID'] = random.randint( 1000, 100000 )
        ## combine all necessary fields to form vector signature
        ## keys-> 'sample_summary'; 'date_range' ; 'hdr_info'

        hdr_info = summary_D['hdr_info']
        sample_summary_ = summary_D['sample_summary']

        unified_key_ =   'Date Range : '+ str( summary_D['date_range'] ) + '\n' \
                       + 'Column Headers : '+ ' , '.join( summary_D['col_names_'] ).strip() + '\n' \
                       + 'LLM Summary : '+ ( sample_summary_ ) if self.add_ai_summary_to_embedding_ is True else ''

        emb_ = createJsonFeats.returnEmbed( unified_key_ )
        insertRec['docSignature'] = emb_
        insertRec['summary'] = unified_key_
        insertRec['file_path'] = summary_D['file_path']
        insertRec['file_name'] = summary_D['file_path'].split('/')[-1]
        insertRec['sheet_name'] = summary_D['sheet_name']
        insertRec['date_range'] = summary_D['date_range']
        insertRec['hdr_info'] = hdr_info

        print('Inserting RECORD->', insertRec['file_name'], insertRec['sheet_name'], unified_key_ )
        return insertRec

    def mergeAndInsert( self, summary_D ):
        '''
        we shall be inserting 2 records for every table
        a) the normal table structure
        b) the transposed table structure
        along with all meta info
        '''
        ##NORM TBL STRUCT
        rec_ = self.createDBRec( summary_D, 'NORM' )
        db_utils.insertNewSignature( rec_ )

    def returnEntireSheet( self, tbl_, sheet_name ):
        '''
        find if the entire sheet contains mostly textual information. If so, then we should simply
        chunk the whole sheet , after concatenating 
        A simple rule of thumb can be the length of the cell contents in any column.
        If the lenght of the cell contents is greater than some threshold say 10 words
        '''
        use_entire_sheet_, chunks_ = False, []

        for col_ctr in range( tbl_['START_COL'], tbl_['END_COL'] ):
            num_str_, unique_, ignore = 0, set(), False
            for row_ctr in range( tbl_['START_ROW'], tbl_['END_ROW'] ):
                if type( self.sheet.cell(row=row_ctr, column=col_ctr).value ) == str and\
                        len( (self.sheet.cell(row=row_ctr, column=col_ctr).value).split() ) >= self.sz_of_phrase_:
                            num_str_ += 1
                unique_.add( ( self.sheet.cell(row=row_ctr, column=col_ctr).value ) )

            ## if num of unique strings in col is low it means, this value is being repeated
            ## HENCE its mostly observations being selected from a drop down and does NOT need
            ## the entire doc chunked
            if len( unique_ ) < self.unique_str_thresh_*( tbl_['END_ROW'] - tbl_['START_ROW'] ): ignore = True
            
            print('returnEntireSheet->', sheet_name, tbl_, num_str_, ( tbl_['END_ROW'] - tbl_['START_ROW'] ), ignore)
            if num_str_ >= self.number_str_thresh_*( tbl_['END_ROW'] - tbl_['START_ROW'] ) and ignore is False:

                use_entire_sheet_ = True
                ## aggregate all text and chunk using self.chunk_size_
                frame_ = ''
                for row_ctr in range( tbl_['START_ROW'], tbl_['END_ROW'] ):
                    for col_ctr in range( tbl_['START_COL'], tbl_['END_COL'] ):
                        
                        if len( frame_ ) >= self.chunk_size_:
                            chunks_.append( frame_ )
                            frame_ = ''

                        frame_ += '\t'+ str( self.sheet.cell(row=row_ctr, column=col_ctr).value )
                    frame_ += '\n'

                if len( frame_ ) > 0: chunks_.append( frame_ )

        return chunks_, use_entire_sheet_

    def process_full_frame_( self, full_frame_, summary_D ):

        for chunk in full_frame_:
           summary_D['sample_summary'] = chunk
           self.mergeAndInsert( summary_D )
        
    def read_excel_file(self):
        # Load the workbook
        main_timer_ = time.time()
        workbook = openpyxl.load_workbook( self.file_path )
        #workbook = openpyxl.load_workbook( self.file_path, read_only=True )
        # Get the specified sheet in the workbook
        summary_D = dict()

        print( ' Time taken to open workbook->', time.time() - main_timer_)
        for sheet_obj in workbook:
            tt_ = time.time()
            self.sheet = sheet_obj
            sheet_name = self.sheet.title
            ## find all tables in the sheet
            #if 'Testing' not in self.sheet.title: continue

            print('Iterating over sheet->', self.sheet.title, self.sheet.max_row)

            all_tables_ = self.find_tables( self.sheet )
            print( 'ALL TABLES in the sheet->', sheet_name, all_tables_)
            print('TIMER: self.find_tables :: ', time.time() - tt_)
            
            for tblidx, tbl_ in enumerate( all_tables_ ):
                frame_, transposed_frame_ = '', ''
                print('xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx')

                if tbl_['START_ROW'] is None or tbl_['END_ROW'] is None or tbl_['START_COL'] is None\
                        or tbl_['END_COL'] is None:
                            print('The tbl and sheet->', self.sheet.title,' no data!')

                try:
                    frame_, high_variance_cols_, col_names_ = self.returnSummary( tbl_ )
                    print('TIMER: self.returnSummary :: ', time.time() - tt_)
                    full_frame_, is_full_frame_necessary_ = self.returnEntireSheet( tbl_, sheet_name )
                    ## find date range if available
                    summary_D['date_range'] = self.findDateRange( tbl_ )
                    summary_D['file_path'] = self.file_path
                    summary_D['sheet_name'] = sheet_name
                    summary_D['col_names_'] = col_names_
                    ## summarize hdr info
                    hdr_frame_ = self.findHeaderInfo( tbl_ )
                    print('TIMER: self.findHeaderInfo :: ', time.time() - tt_)

                    #summary_D['hdr_info'] = self.query_fn_.augmentHeaderInformation( hdr_frame_ )
                    summary_D['hdr_info'] = self.file_path.split('/')[-1] + ' ' + sheet_name + ' ' + \
                                            hdr_frame_ #+ ' ' + summary_D['hdr_info']

                    print('TIMER: self.findDateRange :: ', time.time() - tt_)

                    if summary_D['date_range'] == ( None, None ):
                        ## just add the timestamp of the file ..backup , BUT better than no time dimension
                        summary_D['date_range'] = \
                            datetime.datetime.fromtimestamp( os.path.getmtime(self.file_path) ).strftime('%B %d, %Y')

                    if is_full_frame_necessary_ == True:
                        self.process_full_frame_( full_frame_, summary_D )
                        print('All TEXT ..hence saving chunks!')
                        continue

                    summary_ = self.query_fn_.returnDocSummary( frame_, high_variance_cols_ )

                    print( tblidx,' :: ', tbl_, '::', '\n', frame_, '\n LLAMA3: ', summary_  )
                    print('TIMER: self.query_fn_.returnDocSummary :: ', time.time() - tt_)
                    time.sleep(1) ## groq APIs while testing this were timing out like crazy

                    #print('Sending to LLM for summary->', summary_, '\n', summary_transposed_)
                    ## append file name, sheet name
                    print('Time taken for first 2 LLM calls->', time.time() - tt_)
                    summary_D['sample_summary'] = self.file_path.split('/')[-1] + ' ' + sheet_name + ' ' + summary_

                    print('Time Taken->', time.time() - tt_)
                    print('Time taken for last LLM calls->', time.time() - tt_)

                    #summary_D['pandas_dataframe'] = self.convertToPandas( tbl_ )
                    #print('Time taken for pandas calls->', time.time() - tt_)
                    ## now MERGE all the info and push into vector DB
                    self.mergeAndInsert( summary_D )
                    print('TIMER: self.mergeAndInsert :: ', time.time() - tt_)
                except:
                    print( 'EXCPN-> '+self.file_path + ' ' + sheet_name + ' ' + traceback.format_exc() )
                    continue

                self.masterInfo_[ sheet_name ] = summary_D

if __name__ == '__main__':
    files_ = os.listdir( './DATA/' )

    for file_ in files_:
        try:
            if 'Indexing Process' not in file_: continue

            get_meta_ = GenerateXLMetaInfo( './DATA/' + file_ )
            get_meta_.read_excel_file()
        except:
            print('EXCPN2-> File Loader FAIL = '+'./DATA/' + file_)
            print( traceback.format_exc() )
            continue

    '''

    get_meta_ = GenerateXLMetaInfo( './DATA/Time & Accuracy.xlsx' )
    get_meta_.read_excel_file()

    '''